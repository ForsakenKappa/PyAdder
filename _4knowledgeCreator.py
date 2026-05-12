# _4knowledgeCreator.py
# Модуль формирования правил и знаний
# На основе документации предметной области
# Алфавит задаётся пользователем в виде строки

if __name__ == "__main__":
    print("This script isn't supposed to be executed separately!")

if __name__ == "knowledgeCreator":
    
    import openpyxl as pyxl
    from openpyxl import Workbook
    import os
    import json

    class KnowledgeCreator:
        """
        Модуль формирования правил и знаний.
        Создаёт и инициализирует базу знаний (knowledgeBase.xlsx)
        с правилами:
        - определение алфавита (задаётся пользователем строкой)
        - определение чисел и числовой прямой
        - операция "Слияние"
        - динамическая таблица слияния
        - правила сравнения (ᛃ - одинаково, ᚺᛃ - не одинаково)
        """

        def __init__(self):
            self.kb_path = "./resources/knowledgeBase.xlsx"
            self.config_path = "./resources/alphabet_config.json"
            self.wb = None
            self.ws = None
            self.alphabet = []
            self.zero_symbol = ""  # символ начала отсчёта
            self.unit_symbol = ""   # символ единичного отрезка

        def get_alphabet_from_user(self):
            """Запрашивает у пользователя алфавит в виде строки"""
            print("\n" + "="*50)
            print("НАСТРОЙКА АЛФАВИТА СИСТЕМЫ")
            print("="*50)
            print("\nАлфавит — это упорядоченный набор символов, которые будут")
            print("использоваться в вычислительной системе.")
            print("Символы должны следовать в порядке возрастания (от меньшего к большему).")
            print("\nПримеры ввода:")
            print("  • '0123456789' - цифры от 0 до 9")
            print("  • 'abcdefghij' - буквы от a до j")
            print("  • '🗿️💦🍆✨🤷‍️' - эмодзи (каждый символ отдельно)")
            print("\nПервый символ строки будет точкой отсчёта (ноль).")
            print("Второй символ будет единичным отрезком.\n")
            
            while True:
                alphabet_str = input("Введите строку алфавита: ").strip()
                
                if len(alphabet_str) < 2:
                    print("ОШИБКА: Алфавит должен содержать минимум 2 символа!")
                    continue
                
                # Проверка на уникальность символов
                if len(set(alphabet_str)) != len(alphabet_str):
                    print("ОШИБКА: В строке алфавита не должно быть повторяющихся символов!")
                    continue
                
                # Преобразуем строку в список символов
                self.alphabet = list(alphabet_str)
                self.zero_symbol = self.alphabet[0]
                self.unit_symbol = self.alphabet[1]
                
                # Подтверждение от пользователя
                print("\n" + "-" * 40)
                print(f"Введённый алфавит ({len(self.alphabet)} символов):")
                for i, sym in enumerate(self.alphabet):
                    print(f"  {i}: '{sym}'")
                print("-" * 40)
                
                confirm = input("Подтвердить алфавит? (y/n): ").lower()
                if confirm == 'y':
                    break
                else:
                    print("\nПовторите ввод алфавита.\n")
            
            # Сохраняем конфигурацию алфавита
            self._save_alphabet_config()
            
            print("\n" + "="*50)
            print(f"АЛФАВИТ УСПЕШНО СОЗДАН ({len(self.alphabet)} символов):")
            print(" → ".join(self.alphabet))
            print(f"Точка отсчёта (начало): '{self.zero_symbol}'")
            print(f"Единичный отрезок: '{self.unit_symbol}'")
            print("="*50 + "\n")
            
            return self.alphabet

        def _save_alphabet_config(self):
            """Сохраняет конфигурацию алфавита в JSON файл"""
            config = {
                "alphabet": self.alphabet,
                "alphabet_string": ''.join(self.alphabet),
                "zero_symbol": self.zero_symbol,
                "unit_symbol": self.unit_symbol,
                "size": len(self.alphabet)
            }
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            print(f"[CONFIG] Алфавит сохранён в {self.config_path}")

        def load_alphabet_config(self):
            """Загружает сохранённую конфигурацию алфавита"""
            if os.path.exists(self.config_path):
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                self.alphabet = config["alphabet"]
                self.zero_symbol = config["zero_symbol"]
                self.unit_symbol = config["unit_symbol"]
                return True
            return False

        def _build_merge_table(self):
            """
            Динамически строит таблицу слияния на основе алфавита.
            Правило: a ᚷ b = алфавит[(index(a) + index(b)) mod размер_алфавита]
            """
            n = len(self.alphabet)
            merge_results = {}
            
            for i, a in enumerate(self.alphabet):
                for j, b in enumerate(self.alphabet):
                    result_index = (i + j) % n
                    result = self.alphabet[result_index]
                    merge_results[(a, b)] = result
            
            return merge_results

        def create_knowledge_base(self, use_saved_config=False):
            """
            Создаёт базу знаний с нуля или перезаписывает существующую
            
            Параметры:
            - use_saved_config: если True и есть сохранённый алфавит, использует его
            """
            # Загружаем или запрашиваем алфавит
            if use_saved_config and self.load_alphabet_config():
                print(f"\n[LOAD] Загружен сохранённый алфавит: {''.join(self.alphabet)}")
                if input("Использовать этот алфавит? (y/n): ").lower() != 'y':
                    self.get_alphabet_from_user()
            else:
                self.get_alphabet_from_user()
            
            # Создаём базу знаний
            if os.path.exists(self.kb_path):
                print("Knowledge base already exists. Overwriting...")
            
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "KnowledgeBase"
            
            # Заполнение базы знаний
            self._add_alphabet()
            self._add_number_line_definition()
            self._add_equality_rules()
            self._add_merge_operation_definition()
            self._add_merge_table()
            
            # Сохранение
            self.wb.save(self.kb_path)
            print(f"Knowledge base created successfully at {self.kb_path}")
            
        def _add_alphabet(self):
            """Добавляет символьный алфавит системы (пользовательский)"""
            self.ws["A1"] = "RULE: ALPHABET"
            self.ws["A2"] = "Description: User-defined ordered set of symbols (string input)"
            self.ws["A3"] = "Alphabet string"
            self.ws["B3"] = ''.join(self.alphabet)
            self.ws["A4"] = "Order"
            self.ws["B4"] = "Symbol"
            self.ws["C4"] = "Index"
            
            for idx, symbol in enumerate(self.alphabet, start=5):
                self.ws[f"A{idx}"] = idx - 4  # порядковый номер
                self.ws[f"B{idx}"] = symbol
                self.ws[f"C{idx}"] = idx - 5  # индекс (0-based)
            
            row = len(self.alphabet) + 6
            self.ws[f"A{row}"] = "Zero/Start symbol"
            self.ws[f"B{row}"] = self.zero_symbol
            row += 1
            self.ws[f"A{row}"] = "Unit segment symbol"
            self.ws[f"B{row}"] = self.unit_symbol
            row += 1
            self.ws[f"A{row}"] = "Alphabet size"
            self.ws[f"B{row}"] = len(self.alphabet)
            
        def _add_number_line_definition(self):
            """Добавляет определение числовой прямой"""
            row = len(self.alphabet) + 16
            self.ws[f"A{row}"] = "RULE: NUMBER LINE DEFINITION"
            row += 1
            self.ws[f"A{row}"] = f"1. Числовая прямая имеет точку отсчёта, направление роста и единичный отрезок"
            row += 1
            self.ws[f"A{row}"] = f"2. Точка отсчёта обозначается символом — «{self.zero_symbol}»"
            row += 1
            self.ws[f"A{row}"] = f"3. Единичный отрезок обозначается символом — «{self.unit_symbol}»"
            row += 1
            self.ws[f"A{row}"] = "4. Направление роста — от начала к бесконечности"
            row += 1
            self.ws[f"A{row}"] = "5. Конец — символ, обозначающий наибольшее значение на числовой прямой"
            row += 1
            self.ws[f"A{row}"] = f"6. Наибольшее значение: «{self.alphabet[-1]}»"
            
        def _add_equality_rules(self):
            """Добавляет правила сравнения чисел"""
            row = len(self.alphabet) + 26
            self.ws[f"A{row}"] = "RULE: NUMBER EQUALITY"
            row += 1
            self.ws[f"A{row}"] = "Одинаковость чисел обозначает, занимают ли два числа одно и то же место на числовой прямой"
            row += 1
            self.ws[f"A{row}"] = "Символ одинаковости: ᛃ"
            row += 1
            self.ws[f"A{row}"] = "Символ неодинаковости: ᚺᛃ"
            row += 2
            
            # Примеры
            self.ws[f"A{row}"] = "Примеры:"
            row += 1
            self.ws[f"A{row}"] = f"{self.unit_symbol}ᛃ{self.unit_symbol}"
            row += 1
            self.ws[f"A{row}"] = f"{self.zero_symbol}ᛃ{self.zero_symbol}"
            row += 1
            self.ws[f"A{row}"] = f"{self.zero_symbol}ᚺᛃ{self.unit_symbol}"
            
        def _add_merge_operation_definition(self):
            """Добавляет определение операции «Слияние»"""
            row = len(self.alphabet) + 39
            self.ws[f"A{row}"] = "RULE: MERGE OPERATION (Слияние)"
            row += 1
            self.ws[f"A{row}"] = "Формула: A ᚷ B = алфавит[(index(A) + index(B)) mod size]"
            row += 1
            self.ws[f"A{row}"] = "Алгоритм слияния двух чисел:"
            row += 1
            self.ws[f"A{row}"] = "1. Совершить путь от начала отсчёта до первого числа"
            row += 1
            self.ws[f"A{row}"] = "2. Не возвращаясь в начало, повторить путь, который ведёт от начала отсчёта ко второму числу"
            row += 1
            self.ws[f"A{row}"] = "3. Если в процессе пути доходите до конца числовой прямой — вернуться в начало и пройти оставшиеся шаги"
            row += 1
            self.ws[f"A{row}"] = "4. Точка, в которой окажетесь в итоге — результат слияния"
            row += 2
            
            # Примеры слияния (первые несколько)
            self.ws[f"A{row}"] = "Примеры:"
            row += 1
            if len(self.alphabet) >= 2:
                self.ws[f"A{row}"] = f"{self.zero_symbol}ᚷ{self.zero_symbol}ᛃ{self.zero_symbol}"
                row += 1
                self.ws[f"A{row}"] = f"{self.zero_symbol}ᚷ{self.unit_symbol}ᛃ{self.unit_symbol}"
                row += 1
            if len(self.alphabet) >= 3:
                self.ws[f"A{row}"] = f"{self.unit_symbol}ᚷ{self.unit_symbol}ᛃ{self.alphabet[2]}"
                row += 1
            
            row += 2
            self.ws[f"A{row}"] = "Символ операции слияния: ᚷ"
            
        def _add_merge_table(self):
            """Добавляет динамически построенную таблицу слияния"""
            merge_results = self._build_merge_table()
            
            start_row = len(self.alphabet) + 56
            self.ws[f"A{start_row}"] = "RULE: MERGE TABLE"
            start_row += 1
            self.ws[f"A{start_row}"] = "Formula: result_index = (index(A) + index(B)) mod alphabet_size"
            start_row += 2
            
            # Заголовки
            self.ws[f"A{start_row}"] = "A (first)"
            self.ws[f"B{start_row}"] = "B (second)"
            self.ws[f"C{start_row}"] = "Index(A)"
            self.ws[f"D{start_row}"] = "Index(B)"
            self.ws[f"E{start_row}"] = "Result (A ᚷ B)"
            start_row += 1
            
            # Заполнение таблицы
            alphabet_indices = {sym: idx for idx, sym in enumerate(self.alphabet)}
            
            for (a, b), result in merge_results.items():
                self.ws[f"A{start_row}"] = a
                self.ws[f"B{start_row}"] = b
                self.ws[f"C{start_row}"] = alphabet_indices[a]
                self.ws[f"D{start_row}"] = alphabet_indices[b]
                self.ws[f"E{start_row}"] = result
                start_row += 1

        def get_merge_result(self, a, b):
            """
            Возвращает результат слияния двух символов
            На основе динамической формулы
            """
            if a not in self.alphabet or b not in self.alphabet:
                return None
            
            idx_a = self.alphabet.index(a)
            idx_b = self.alphabet.index(b)
            result_idx = (idx_a + idx_b) % len(self.alphabet)
            return self.alphabet[result_idx]

        def is_valid_symbol(self, symbol):
            """Проверяет, является ли символ допустимым"""
            return symbol in self.alphabet

        def get_alphabet(self):
            """Возвращает текущий алфавит"""
            return self.alphabet.copy()
        
        def get_alphabet_string(self):
            """Возвращает алфавит в виде строки"""
            return ''.join(self.alphabet)

    # Автоматическое создание базы знаний при импорте модуля
    _creator = KnowledgeCreator()
    
    # Проверяем, нужно ли создать новую БЗ или использовать существующую
    if os.path.exists("./resources/knowledgeBase.xlsx") and os.path.exists("./resources/alphabet_config.json"):
        print("\n[INFO] Существующая база знаний найдена.")
        response = input("Создать новую базу знаний с другим алфавитом? (y/n): ").lower()
        if response == 'y':
            _creator.create_knowledge_base(use_saved_config=False)
        else:
            _creator.load_alphabet_config()
            print(f"[INFO] Используется сохранённый алфавит: {_creator.get_alphabet_string()}")
    else:
        _creator.create_knowledge_base(use_saved_config=False)
